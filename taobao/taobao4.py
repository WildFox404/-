import time
import json
import openpyxl
from DrissionPage import WebPage
import re
class TaoBao():
    def __init__(self,workbook):
        self.page=WebPage()
        self.sheet=workbook["Sheet1"]
    def login(self):
        self.page.get("https://login.taobao.com/member/login.jhtml")
        time.sleep(10)
        self.page.ele("@class=search-combobox-input",timeout=20).input("abc")

    def get_data(self):
        data=self.page.listen.wait(timeout=2)
        if data==False:
        # if data.response.body==None:
            print("未获取到数据")
            return""
        else:
            return data.response.body
    def search_data(self,data):
        try:
            result_list=[]
            # 查找JSON数据起始位置
            start_index = data.find("{")
            # 查找JSON数据终止位置
            end_index = data.rfind("}") + 1

            # 如果找到起始和终止位置，则提取JSON数据
            if start_index != -1 and end_index != -1:
                json_data = data[start_index:end_index]
            else:
                print("未找到JSON数据")
            json_data = json_data.encode('utf-8')
            json_data = json.loads(json_data)
            data_list = json_data['data']['data']['items']

            for i in range(len(data_list)):
                sell_fuzzy_num = int(re.search(r'\d+', data_list[i]['material']['sellFuzzy']).group())
                if sell_fuzzy_num < 1:
                    print(data_list[i]['material']['eurl'])  # 店铺链接
                    print(data_list[i]['material']['sellFuzzy'])  # 销售额
                    result_list.append(data_list[i]['material']['eurl'])
            return result_list
        except:
            print("json数据出错")
            time.sleep(5)
            return []

def get_data(page):
    data=page.listen.wait(timeout=3)
    if data.response.body==None:
        return False
    else:
        return True


def get_shop_url(data):
    # 使用正则表达式匹配
    pattern = r"'pcShopUrl':\s*'(.*?)'"
    result = re.search(pattern, data)

    # 提取匹配到的内容
    if result:
        extracted_content = result.group(1)
        print(extracted_content)  # 输出：//shop179492430.taobao.com
        return extracted_content
    else:
        print("未匹配到内容")
        return ""
def get_data_result(page):
    data = page.listen.wait(timeout=2)
    if data.response.body == None:
        print("未获取到数据")
    else:
        return data.response.body
if __name__ == '__main__':
    count=0
    wb = openpyxl.load_workbook("taobao2.xlsx")
    taobao=TaoBao(wb)
    taobao.login()
    taobao.page.listen.start("h5/mtop.taobao.pcdetail.data.get/1.0/")
    # Iterate over the rows in the first column
    if __name__ == '__main__':
        count = 0
        wb = openpyxl.load_workbook("taobao3.xlsx")
        taobao = TaoBao(wb)
        taobao.login()
        taobao.page.listen.start("h5/mtop.taobao.pcdetail.data.get/1.0/")
        # Iterate over the rows in the first column
        for row_index, cell in enumerate(taobao.sheet['A'], start=1):  # start index from 1
            time.sleep(3)
            # Access cell value
            print(cell.value)
            taobao.page.get(cell.value)
            data = taobao.get_data()
            print(str(data))
            result = get_shop_url(str(data))

            # Write data into the second column of the same row
            if result == "":
                print("未获取到店铺链接")
                taobao.sheet.cell(row=row_index, column=2, value="未找到链接")
            else:
                print(result)
                # Write 'result' into the corresponding cell in the second column
                taobao.sheet.cell(row=row_index, column=2, value=result)
            wb.save("taobao3.xlsx")

            # Save the workbook

                # for i in range(100):
    #     time.sleep(1)
    #     taobao.page.get(f"https://uland.taobao.com/sem/tbsearch?keyword=%E6%89%8B%E6%9C%BA%E5%BE%AE%E4%B9%90%E5%B0%8F%E7%A8%8B%E5%BA%8F%E5%AE%B6%E4%B9%A1%E5%B9%BF%E4%B8%9C%E9%BA%BB%E5%B0%86&pnum={i+1}")
    #     data=taobao.get_data()
    #     if data=="":
    #         time.sleep(2)
    #         continue
    #     print("第{}页数据获取成功\n".format(i+1))
    #     result_list = taobao.search_data(data)
    #     for eurl in result_list:
    #         result=[]
    #         result.append(eurl)
    #         taobao.sheet.append(result)
    #         count+=1
    #     print("第{}页数据写入成功\n".format(i+1))
    #     print("当前数据量:"+count)
    #     wb.save("taobao2.xlsx")